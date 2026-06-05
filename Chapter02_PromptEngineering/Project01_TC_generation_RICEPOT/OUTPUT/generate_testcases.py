"""Generate restful-booker API test cases as XLSX per RICE-POT prompt rules.
All test cases derived strictly from the attached restful-booker API PDF.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

HEADERS = [
    "Scenario", "TID", "Test Data", "Test Case Description", "Pre-Condition",
    "Test Steps", "Expected Result", "Actual Result", "Status",
    "Executed By (QA Name)", "Misc (Comments)", "Priority", "Is Automated",
]

ROWS = [
    # ---------------- Auth - CreateToken ----------------
    ["Auth - CreateToken (Positive)", "TC-001",
     'username="admin", password="password123"',
     "Verify a valid auth token is returned when correct default credentials are POSTed to /auth.",
     "API is reachable at https://restful-booker.herokuapp.com",
     "1. POST https://restful-booker.herokuapp.com/auth | 2. Set header Content-Type: application/json | 3. Send body {\"username\":\"admin\",\"password\":\"password123\"}",
     "HTTP 200 OK with JSON body containing a string field \"token\" (e.g. {\"token\":\"abc123\"}).",
     "", "", "", "Derived from PDF: Auth - CreateToken", "High", "Yes"],

    ["Auth - CreateToken (Negative - invalid password)", "TC-002",
     'username="admin", password="wrongPass"',
     "Verify CreateToken does not issue a valid token when password is invalid.",
     "API is reachable",
     "1. POST /auth with Content-Type: application/json | 2. Body {\"username\":\"admin\",\"password\":\"wrongPass\"}",
     "Response must NOT contain a valid \"token\" field. Inference (low confidence): exact error payload not specified in PDF.",
     "", "", "", "PDF defines only success response; failure payload not specified.", "High", "Yes"],

    ["Auth - CreateToken (Negative - missing body)", "TC-003",
     "empty request body",
     "Verify CreateToken behavior when request body is empty.",
     "API is reachable",
     "1. POST /auth with Content-Type: application/json | 2. Send empty body",
     "Insufficient information to determine.",
     "", "", "", "PDF does not document behavior for empty body.", "Medium", "No"],

    # ---------------- Booking - GetBookingIds ----------------
    ["Booking - GetBookingIds (All)", "TC-004",
     "no query parameters",
     "Verify GET /booking returns an array of objects containing unique bookingid values.",
     "At least one booking exists in the system",
     "1. GET https://restful-booker.herokuapp.com/booking",
     "HTTP 200 OK with JSON array; each element is an object with numeric field \"bookingid\".",
     "", "", "", "Derived from PDF: Booking - GetBookingIds Example 1", "High", "Yes"],

    ["Booking - GetBookingIds (Filter by firstname & lastname)", "TC-005",
     "firstname=sally, lastname=brown",
     "Verify filtering booking IDs by firstname and lastname query parameters returns only matching bookings.",
     "A booking with firstname=sally and lastname=brown exists",
     "1. GET /booking?firstname=sally&lastname=brown",
     "HTTP 200 OK with JSON array of objects each containing only a numeric \"bookingid\" for bookings matching firstname=sally AND lastname=brown.",
     "", "", "", "Derived from PDF: Example 2 (Filter by name)", "High", "Yes"],

    ["Booking - GetBookingIds (Filter by checkin & checkout)", "TC-006",
     "checkin=2014-03-13, checkout=2014-05-21",
     "Verify filtering by checkin/checkout dates (CCYY-MM-DD) returns bookings with dates greater than or equal to the supplied values.",
     "Bookings exist with the relevant date ranges",
     "1. GET /booking?checkin=2014-03-13&checkout=2014-05-21",
     "HTTP 200 OK with JSON array of bookingid objects where checkin >= 2014-03-13 AND checkout >= 2014-05-21.",
     "", "", "", "Derived from PDF: Example 3 (Filter by checkin/checkout date)", "High", "Yes"],

    ["Booking - GetBookingIds (Negative - invalid date format)", "TC-007",
     "checkin=13-03-2014 (DD-MM-YYYY)",
     "Verify behavior when checkin is supplied in a non-CCYY-MM-DD format.",
     "API is reachable",
     "1. GET /booking?checkin=13-03-2014",
     "Insufficient information to determine.",
     "", "", "", "PDF specifies format must be CCYY-MM-DD but does not document error response.", "Medium", "No"],

    # ---------------- Booking - GetBooking ----------------
    ["Booking - GetBooking (JSON)", "TC-008",
     "id=1, Accept: application/json",
     "Verify GET /booking/:id returns full booking details in JSON when Accept header is application/json.",
     "Booking with id=1 exists",
     "1. GET /booking/1 | 2. Header Accept: application/json",
     "HTTP 200 OK with JSON body containing firstname (String), lastname (String), totalprice (Number), depositpaid (Boolean), bookingdates.checkin (Date), bookingdates.checkout (Date), additionalneeds (String).",
     "", "", "", "Derived from PDF: Booking - GetBooking JSON Response", "High", "Yes"],

    ["Booking - GetBooking (XML)", "TC-009",
     "id=1, Accept: application/xml",
     "Verify GET /booking/:id returns booking details in XML when Accept header is application/xml.",
     "Booking with id=1 exists",
     "1. GET /booking/1 | 2. Header Accept: application/xml",
     "HTTP 200 OK with XML body rooted at <booking> containing firstname, lastname, totalprice, depositpaid, bookingdates (checkin, checkout) and additionalneeds elements.",
     "", "", "", "Derived from PDF: Booking - GetBooking XML Response", "Medium", "Yes"],

    ["Booking - GetBooking (Default Accept)", "TC-010",
     "id=1, no Accept header",
     "Verify default Accept header is application/json when none is supplied.",
     "Booking with id=1 exists",
     "1. GET /booking/1 with no Accept header",
     "HTTP 200 OK with JSON body (default Accept value: application/json as documented).",
     "", "", "", "Derived from PDF: Accept header Default value application/json", "Medium", "Yes"],

    # ---------------- Booking - CreateBooking ----------------
    ["Booking - CreateBooking (JSON Positive)", "TC-011",
     'firstname="Jim", lastname="Brown", totalprice=111, depositpaid=true, bookingdates.checkin="2018-01-01", bookingdates.checkout="2019-01-01", additionalneeds="Breakfast"',
     "Verify a new booking is created via POST /booking with a JSON payload.",
     "API is reachable",
     "1. POST /booking | 2. Header Content-Type: application/json | 3. Body: {firstname,Jim; lastname,Brown; totalprice,111; depositpaid,true; bookingdates:{checkin,2018-01-01; checkout,2019-01-01}; additionalneeds,Breakfast}",
     "HTTP 200 OK with JSON body containing numeric \"bookingid\" and a \"booking\" object echoing all submitted fields exactly.",
     "", "", "", "Derived from PDF: Booking - CreateBooking JSON example & response", "High", "Yes"],

    ["Booking - CreateBooking (XML Positive)", "TC-012",
     "<booking> with firstname=Jim, lastname=Brown, totalprice=111, depositpaid=true, bookingdates.checkin=2018-01-01, bookingdates.checkout=2019-01-01, additionalneeds=Breakfast",
     "Verify a new booking is created via POST /booking with an XML payload.",
     "API is reachable",
     "1. POST /booking | 2. Header Content-Type: text/xml | 3. Send the XML <booking> payload",
     "HTTP 200 OK with XML body rooted at <created-booking> containing <bookingid> and a <booking> element with all submitted fields.",
     "", "", "", "Derived from PDF: CreateBooking XML example & XML Response", "Medium", "Yes"],

    ["Booking - CreateBooking (URL-encoded Positive)", "TC-013",
     "firstname=Jim&lastname=Brown&totalprice=111&depositpaid=true&bookingdates[checkin]=2018-01-01&bookingdates[checkout]=2019-01-01&additionalneeds=Breakfast",
     "Verify a new booking is created via POST /booking with an application/x-www-form-urlencoded payload.",
     "API is reachable",
     "1. POST /booking | 2. Header Content-Type: application/x-www-form-urlencoded | 3. Send URL-encoded body",
     "HTTP 200 OK with body echoing the booking fields in URL-encoded form (as shown in PDF URL Response).",
     "", "", "", "Derived from PDF: CreateBooking URLencoded example & URL Response", "Medium", "No"],

    ["Booking - CreateBooking (Negative - missing required field)", "TC-014",
     "payload missing firstname",
     "Verify behavior of POST /booking when a documented request body field (firstname) is omitted.",
     "API is reachable",
     "1. POST /booking | 2. Content-Type: application/json | 3. Body without firstname",
     "Insufficient information to determine.",
     "", "", "", "PDF lists firstname under Request body but does not document required/optional status or error payload.", "High", "No"],

    # ---------------- Booking - UpdateBooking ----------------
    ["Booking - UpdateBooking (Cookie token, JSON)", "TC-015",
     'id=1, Cookie: token=abc123, body firstname="James", lastname="Brown", totalprice=111, depositpaid=true, checkin=2018-01-01, checkout=2019-01-01, additionalneeds="Breakfast"',
     "Verify PUT /booking/:id updates a booking when authorized via Cookie token header.",
     "Booking with id=1 exists; valid token obtained from /auth",
     "1. PUT /booking/1 | 2. Headers Content-Type: application/json, Accept: application/json, Cookie: token=abc123 | 3. Send full JSON body",
     "HTTP 200 OK with JSON body reflecting the updated booking fields exactly as submitted.",
     "", "", "", "Derived from PDF: UpdateBooking JSON example & response", "High", "Yes"],

    ["Booking - UpdateBooking (Basic auth, XML)", "TC-016",
     "id=1, Authorization: Basic YWRtaW46cGFzc3dvcmQxMjM=, XML body",
     "Verify PUT /booking/:id updates a booking when authorized via Basic auth header and XML payload.",
     "Booking with id=1 exists",
     "1. PUT /booking/1 | 2. Headers Content-Type: text/xml, Accept: application/xml, Authorization: Basic YWRtaW46cGFzc3dvcmQxMjM= | 3. Send <booking> XML payload",
     "HTTP 200 OK with XML <booking> body reflecting the updated fields.",
     "", "", "", "Derived from PDF: UpdateBooking XML example & response", "Medium", "Yes"],

    ["Booking - UpdateBooking (Negative - missing auth)", "TC-017",
     "id=1, no Cookie and no Authorization header",
     "Verify PUT /booking/:id is rejected when neither Cookie token nor Basic Authorization header is supplied.",
     "Booking with id=1 exists",
     "1. PUT /booking/1 | 2. Headers Content-Type: application/json (no auth headers) | 3. Send full JSON body",
     "Request must be rejected (booking must not be updated). Inference (low confidence): exact status code not specified in PDF.",
     "", "", "", "PDF states PUT requires Cookie or Authorization but does not document the rejection response code.", "High", "Yes"],

    # ---------------- Booking - PartialUpdateBooking ----------------
    ["Booking - PartialUpdateBooking (JSON)", "TC-018",
     'id=1, Cookie: token=abc123, body {"firstname":"James","lastname":"Brown"}',
     "Verify partial update of a booking with only firstname and lastname supplied.",
     "Booking with id=1 exists; valid token",
     "1. PATCH /booking/1 | 2. Headers Content-Type: application/json, Accept: application/json, Cookie: token=abc123 | 3. Body {firstname,James; lastname,Brown}",
     "HTTP 200 OK with full booking JSON reflecting the updated firstname/lastname and retaining other fields.",
     "", "", "", "Derived from PDF: PartialUpdateBooking JSON example & response. Note: PDF documents method as PATCH but cURL examples use PUT.", "High", "Yes"],

    ["Booking - PartialUpdateBooking (Negative - no auth)", "TC-019",
     "id=1, no auth headers, partial body",
     "Verify PATCH /booking/:id is rejected without Cookie or Authorization header.",
     "Booking with id=1 exists",
     "1. PATCH /booking/1 | 2. Content-Type: application/json (no auth) | 3. Body {firstname,James}",
     "Request must be rejected; booking must remain unchanged. Inference (low confidence): exact status code not specified.",
     "", "", "", "PDF requires auth header (same pattern as Update).", "High", "No"],

    # ---------------- Booking - DeleteBooking ----------------
    ["Booking - DeleteBooking (Cookie token)", "TC-020",
     "id=1, Cookie: token=abc123",
     "Verify DELETE /booking/:id removes a booking when authorized via Cookie token.",
     "Booking with id=1 exists; valid token",
     "1. DELETE /booking/1 | 2. Headers Content-Type: application/json, Cookie: token=abc123",
     "HTTP 201 Created (per PDF: Success 200 'OK' field, Default HTTP 201 response).",
     "", "", "", "Derived from PDF: DeleteBooking Example 1 & Success response", "High", "Yes"],

    ["Booking - DeleteBooking (Basic auth)", "TC-021",
     "id=1, Authorization: Basic YWRtaW46cGFzc3dvcmQxMjM=",
     "Verify DELETE /booking/:id removes a booking when authorized via Basic auth header.",
     "Booking with id=1 exists",
     "1. DELETE /booking/1 | 2. Headers Content-Type: application/json, Authorization: Basic YWRtaW46cGFzc3dvcmQxMjM=",
     "HTTP 201 Created (per PDF Success response).",
     "", "", "", "Derived from PDF: DeleteBooking Example 2", "High", "Yes"],

    ["Booking - DeleteBooking (Negative - no auth)", "TC-022",
     "id=1, no auth headers",
     "Verify DELETE /booking/:id is rejected without Cookie or Authorization header.",
     "Booking with id=1 exists",
     "1. DELETE /booking/1 | 2. Header Content-Type: application/json only",
     "Request must be rejected; booking must remain. Inference (low confidence): exact status code not specified.",
     "", "", "", "PDF: DELETE requires authorization token or Basic auth header.", "High", "Yes"],

    # ---------------- Ping - HealthCheck ----------------
    ["Ping - HealthCheck", "TC-023",
     "no parameters",
     "Verify the health check endpoint returns HTTP 201 Created when the API is up.",
     "API is reachable",
     "1. GET https://restful-booker.herokuapp.com/ping",
     "HTTP 201 Created (per PDF Ping - HealthCheck response).",
     "", "", "", "Derived from PDF: Ping - HealthCheck", "High", "Yes"],

    # ---------------- Non-functional ----------------
    ["Non-Functional - Content negotiation (XML on GetBooking)", "TC-024",
     "id=1, Accept: application/xml",
     "Verify content negotiation: server returns XML when Accept: application/xml is requested.",
     "Booking with id=1 exists",
     "1. GET /booking/1 | 2. Header Accept: application/xml",
     "Response Content-Type matches application/xml and body is valid XML rooted at <booking>.",
     "", "", "", "Derived from PDF: GetBooking Accept header description", "Medium", "Yes"],

    ["Non-Functional - Default Content-Type on CreateBooking", "TC-025",
     "POST /booking without explicit Content-Type",
     "Verify documented default Content-Type (application/json) is applied when client omits the header.",
     "API is reachable",
     "1. POST /booking with valid JSON body and no Content-Type header",
     "Insufficient information to determine.",
     "", "", "", "PDF states Default value: application/json but does not specify server behavior when header is omitted.", "Low", "No"],
]


def main() -> None:
    out_dir = Path(__file__).resolve().parent
    out_path = out_dir / "RestfulBooker_TestCases.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ROWS:
        ws.append(row)

    widths = [38, 10, 38, 50, 38, 60, 60, 14, 12, 22, 40, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"Wrote {out_path} with {len(ROWS)} test cases.")


if __name__ == "__main__":
    main()
